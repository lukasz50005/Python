#!/usr/bin/python3
import smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders
import re
from openpyxl import load_workbook
from openpyxl import Workbook
import psycopg2

#dane do pliku: ile jest pobieranych kolumn oraz wierszy
ileKolumn_plik = 4
maksymalnie_wierszy_plik=100

#dane do bazy danych
db="postgres"
dbuser="postgres"
dbpass="postgres"
dbhost="localhost"
dbport="5432"

#wysyłanie wiadomości
def send_mail( send_from, send_to, subject, text, files=[], server="localhost", port=587, username='', password='', isTls=True):
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime = True)
    msg['Subject'] = subject

    msg.attach( MIMEText(text) )
    for f in files:
        part = MIMEBase('application', "octet-stream")
        part.set_payload( open(f,"rb").read() )
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename="{0}"'.format(os.path.basename(f)))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    if isTls: smtp.starttls()
    smtp.login(username,password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.quit()
    
#sprawdzanie czy adres e-mail jest poprawny
def adres(mail):
	return re.match( r'([\w\.\+\-]+)@[\w]+\.\w{2,4}$', mail)

#odczyt z pliku i wstawianie danych do bazy    
def baza(ws,ileKolumn,maksymalnie_wierszy,db,dbuser,dbpass,dbhost,dbport):
    i = 0
    blad = 0
    poprawnie = 1
    daneKolumn = []
    wiersz = []
    for row in ws.iter_rows( max_col=ileKolumn, max_row=maksymalnie_wierszy):
        for cell in row:
            if(str(cell.value) != 'None'):
                wiersz.append(str(cell.value))
        if(len(wiersz) > 0):
            daneKolumn.append(wiersz)
            wiersz = []
        
    try:
        conn = psycopg2.connect(database=db, user=dbuser, password=dbpass, host=dbhost, port=dbport)
        cur = conn.cursor()
    except:
      print ("\n--------Nie można połączyć się z bazą danych.--------\n")
      poprawnie = 0
    if (poprawnie == 1):  
        try:
          cur.execute("CREATE TABLE IF NOT EXISTS public.dane(id int not null,imie char(30) not null,nazwisko char(50),wiek int not null,miasto char(50));")
        except:
          print ("\nNie mogę stworzyć tabeli\n")
          poprawnie = 0
        try:
          cur.execute("select id from public.dane order by id desc limit 1;")
          id_row = cur.fetchall()
          if (len(id_row)!= 0):
            id = id_row[0][0]
          else:
            id = 0
        except:
          poprawnie = 0
        for dane in daneKolumn:    
            if(i != 0 and len(dane)==4):
                try:
                  id+=1
                  cur.execute("insert into dane(id,imie,nazwisko,wiek,miasto) values ("+str(id)+",'"+dane[0]+"', '"+dane[1]+"', "+dane[2]+",'"+dane[3]+"');")
                except:
                  blad+=1
                  if(blad == 1):
                    print ("\nPlik nie został dodany do bazy.\nBłąd danych w :"+dane[0]+"', '"+dane[1]+"', "+dane[2]+",'"+dane[3]+'\n')
                    poprawnie = 0
            i+=1
        try:
            conn.commit()
            cur.close()
            conn.close()
        except:
            print ("\nNie mogę zamknąć bazy.\n")
            poprawnie = 0
    if(poprawnie == 1):
        print("Dane zapisane w bazie.")
        
#informacje do wysłania wiadomości
while True:
    email_od=input('\nPodaj twój adres e-mail\n')
    mail_1 = adres(email_od)
    if (mail_1):
        break
    else:
        print('\n\n--------To nie jest adres email--------')
        
while True:
    email_do=input('\nDo kogo chcesz wysłać?\n')
    mail_2 = adres(email_do)
    if (mail_2):
        break
    else:
        print('\n\n--------To nie jest adres email--------')

temat=input('\nWpisz temat wiadomości\n')
tresc=input('\nWpisz tresc wiadomości\n')
files=[]
while True:
    lokalizacja = input('\nPodaj lokalizacje załącznika lub 0 aby nie dodawać pliku.\n')
    if (lokalizacja == '0'):
        break
    else:
        try:
            at = open(lokalizacja, "r")
            at.close()
            files.append(lokalizacja)
            break
        except:
            print('\n--------Plik w podanej lokalizacji nie istnieje.--------\n')
            
serw=input('\nPodaj adres serwera smtp\n')
haslo=input('\nWpisz swoje hasło do poczty\n')


print ('Wysyłanie wiadomości...\n')

try:
    send_mail( email_od, [email_do], temat, tresc, files, server=serw, port=587, username=email_od, password=haslo, isTls=True)
    
    print ('Wysłano wiadomość\n')
    
    nazwaPliku = re.search( r'dane\.xlsx$', lokalizacja, re.M|re.I)

    if nazwaPliku:
        wb = load_workbook(lokalizacja)
        ws = wb.active
        baza(ws,ileKolumn_plik,maksymalnie_wierszy_plik,db,dbuser,dbpass,dbhost,dbport)
except:
    print ('\n--------Nie można wysłać wiadomości.--------\n--------Sprawdź czy dane są poprawne.--------\n')
    
print  (" Copyright Łukasz Grześkowiak ")
    



